import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from utils.logger import logger


class ExcelUtils:
    """Excel操作工具类"""

    @staticmethod
    def read_config_sheet(excel_path):
        """读取配置工作表"""
        try:
            wb = load_workbook(excel_path)
            logger.debug(f"Excel文件中的所有工作表: {wb.sheetnames}")

            if 'config' not in wb.sheetnames:
                logger.warning("Excel中未找到config工作表，使用默认配置")
                return {}

            ws = wb['config']
            config = {}

            # 假设第一行为表头：key, value
            for row in range(2, ws.max_row + 1):
                key_cell = ws.cell(row=row, column=1)
                value_cell = ws.cell(row=row, column=2)
                if key_cell.value is None:
                    continue
                config[key_cell.value] = value_cell.value
            return config
        except Exception as e:
            logger.error(f"读取config工作表失败: {str(e)}")
            return {}

    @staticmethod
    def read_interfaces_sheet(excel_path):
        """读取接口工作表"""
        try:
            wb = load_workbook(excel_path)
            if 'interfaces' not in wb.sheetnames:
                logger.error("Excel中未找到interfaces工作表")
                return {}

            ws = wb['interfaces']
            interfaces = {}

            # 假设表头行：接口ID, 接口名称, 接口URL, 请求方法, 请求头, Content-Type
            header_row = 1
            headers = {}
            for col in range(1, ws.max_column + 1):
                headers[ws.cell(row=header_row, column=col).value] = col

            # 检查必要的表头是否存在
            required_headers = ['接口ID', '接口名称', '接口URL', '请求方法']
            for header in required_headers:
                if header not in headers:
                    logger.error(f"interfaces工作表缺少必要的表头: {header}")
                    return {}

            # 读取接口数据
            for row in range(2, ws.max_row + 1):
                interface_id = ws.cell(row=row, column=headers['接口ID']).value
                if not interface_id:
                    continue

                # 解析请求头（JSON字符串）
                headers_str = ws.cell(row=row, column=headers.get('请求头', 0)).value
                headers_data = {}
                if headers_str:
                    try:
                        headers_data = json.loads(headers_str)
                    except json.JSONDecodeError:
                        logger.warning(f"接口 {interface_id} 的请求头格式不正确")

                interfaces[interface_id] = {
                    'name': ws.cell(row=row, column=headers['接口名称']).value,
                    'url': ws.cell(row=row, column=headers['接口URL']).value,
                    'method': ws.cell(row=row, column=headers['请求方法']).value or 'GET',
                    'headers': headers_data,
                    'content_type': ws.cell(row=row, column=headers.get('Content-Type', 0)).value or 'application/json'
                }

            return interfaces
        except Exception as e:
            logger.error(f"读取interfaces工作表失败: {str(e)}")
            return {}

    @staticmethod
    def read_testcases_sheet(excel_path):
        """读取测试用例工作表"""
        try:
            wb = load_workbook(excel_path)
            if 'testcases' not in wb.sheetnames:
                logger.error("Excel中未找到testcases工作表")
                return []

            ws = wb['testcases']
            testcases = []

            # 读取表头
            header_row = 1
            headers = {}
            for col in range(1, ws.max_column + 1):
                header_name = ws.cell(row=header_row, column=col).value
                if header_name:
                    headers[header_name] = col

            # 检查必要的表头是否存在
            required_headers = ['用例ID', '用例名称', '接口ID']
            for header in required_headers:
                if header not in headers:
                    logger.error(f"testcases工作表缺少必要的表头: {header}")
                    return []

            # 检查是否有数据行
            if ws.max_row < 2:
                logger.warning("testcases工作表中没有测试用例数据")
                return []

            # 读取用例数据
            for row in range(2, ws.max_row + 1):
                case_id = ws.cell(row=row, column=headers['用例ID']).value
                if not case_id:
                    continue

                # 解析请求参数（JSON字符串）
                params_col = headers.get('请求参数')
                params_str = None
                if params_col and params_col > 0:
                    params_str = ws.cell(row=row, column=params_col).value
                params_data = {}
                if params_str:
                    try:
                        params_data = json.loads(params_str)
                    except json.JSONDecodeError:
                        logger.warning(f"用例 {case_id} 的请求参数格式不正确")

                # 解析请求头（JSON字符串）
                headers_col = headers.get('请求头')
                headers_str = None
                if headers_col and headers_col > 0:
                    headers_str = ws.cell(row=row, column=headers_col).value
                headers_data = {}
                if headers_str:
                    try:
                        headers_data = json.loads(headers_str)
                    except json.JSONDecodeError:
                        logger.warning(f"用例 {case_id} 的请求头格式不正确")

                # 解析提取变量规则（JSON字符串）
                extract_col = headers.get('提取变量')
                extract_str = None
                if extract_col and extract_col > 0:
                    extract_str = ws.cell(row=row, column=extract_col).value
                extract_data = {}
                if extract_str:
                    try:
                        extract_data = json.loads(extract_str)
                    except json.JSONDecodeError:
                        logger.warning(f"用例 {case_id} 的提取变量格式不正确")

                # 解析断言规则（JSON字符串）
                assert_col = headers.get('断言规则')
                assert_str = None
                if assert_col and assert_col > 0:
                    assert_str = ws.cell(row=row, column=assert_col).value
                assert_data = {}
                if assert_str:
                    try:
                        assert_data = json.loads(assert_str)
                    except json.JSONDecodeError:
                        logger.warning(f"用例 {case_id} 的断言规则格式不正确")

                # 处理其他字段
                module_col = headers.get('模块')
                module_value = None
                if module_col and module_col > 0:
                    module_value = ws.cell(row=row, column=module_col).value

                dependent_col = headers.get('依赖用例ID')
                dependent_value = None
                if dependent_col and dependent_col > 0:
                    dependent_value = ws.cell(row=row, column=dependent_col).value

                run_col = headers.get('是否运行')
                run_value = 'YES'
                if run_col and run_col > 0:
                    run_value = ws.cell(row=row, column=run_col).value or 'YES'

                testcase = {
                    '用例ID': case_id,
                    '用例名称': ws.cell(row=row, column=headers['用例名称']).value,
                    '接口ID': ws.cell(row=row, column=headers['接口ID']).value,
                    '模块': module_value,
                    '请求参数': params_data,
                    '请求头': headers_data,
                    '依赖用例ID': dependent_value,
                    '提取变量': extract_data,
                    '断言规则': assert_data,
                    '是否运行': run_value
                }

                testcases.append(testcase)

            return testcases
        except Exception as e:
            logger.error(f"读取testcases工作表失败: {str(e)}")
            return []

    @staticmethod
    def save_test_results(excel_path, results):
        """保存测试结果到Excel"""
        try:
            # 加载工作簿
            try:
                wb = load_workbook(excel_path)
            except FileNotFoundError:
                wb = Workbook()

            # 如果存在结果表则删除
            if 'testresults' in wb.sheetnames:
                del wb['testresults']

            # 创建结果表
            ws = wb.create_sheet('testresults', 3)  # 插入到第4个位置
            ws.title = 'testresults'

            # 写入表头
            headers = ['用例ID', '用例名称', '接口ID', '开始时间', '结束时间',
                       '响应时间(ms)', '状态码', '测试结果', '错误信息', '断言结果']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.alignment = Alignment(wrap_text=True)
                ws.column_dimensions[chr(64 + col)].width = 20  # 设置列宽

            # 调整部分列宽
            ws.column_dimensions['I'].width = 30  # 错误信息
            ws.column_dimensions['J'].width = 50  # 断言结果

            # 写入结果数据
            for row_idx, result in enumerate(results, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=result.get(header, ''))
                    cell.alignment = Alignment(wrap_text=True)

                # 为失败的用例标记颜色
                if result.get('测试结果') == '失败':
                    for col_idx in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=col_idx).font = cell.font.copy(color="FF0000")

            # 保存文件
            wb.save(excel_path)
            logger.info(f"测试结果已保存到 {excel_path} 的 testresults 工作表")

        except Exception as e:
            logger.error(f"保存测试结果失败: {str(e)}")
